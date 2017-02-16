# coding=utf-8
# create by oldman at 17/2/16
import tablib
from openpyxl import load_workbook


def deaul_refund_anytime():
    list_result = []
    wb1 = load_workbook(filename='/Users/lipeng/Desktop/doctors/1.xlsx')
    sheetnames1 = wb1.get_sheet_names()
    if sheetnames1 == 0:
        return
    ws1 = wb1.get_sheet_by_name(sheetnames1[0])
    for rx1 in range(2, ws1.get_highest_row() + 1):
        try:
            user_id = int(ws1.cell(row=rx1, column=1).value)
            list_result.append(user_id)
        except Exception as e:
            print(e)
            continue

    wb2 = load_workbook(filename='/Users/lipeng/Desktop/doctors/2.xlsx')
    sheetnames2 = wb2.get_sheet_names()
    if sheetnames2 == 0:
        return
    ws2 = wb2.get_sheet_by_name(sheetnames2[0])
    for rx2 in range(2, ws2.get_highest_row() + 1):
        try:
            user_id = int(ws2.cell(row=rx2, column=1).value)
            list_result.append(user_id)
        except Exception as e:
            print(e)
            continue

    wb3 = load_workbook(filename='/Users/lipeng/Desktop/doctors/3.xlsx')
    sheetnames3 = wb3.get_sheet_names()
    if sheetnames3 == 0:
        return
    ws3 = wb3.get_sheet_by_name(sheetnames3[0])
    for rx3 in range(2, ws3.get_highest_row() + 1):
        try:
            user_id = int(ws3.cell(row=rx3, column=1).value)
            list_result.append(user_id)
        except Exception as e:
            print(e)
            continue
    print(len(list_result))
    print(len(set(list_result)))

    file = open('/Users/lipeng/Desktop/doctors/result.txt', 'a')
    for item in set(list_result):
        file.write(str(item)+'\n')

    file.close()

    # print(dict_service)
    # print([values for key, values in dict_service.items() if values == 'False'])
    # dt_service_refund_anytime = tablib.Dataset()
    # dt_service_refund_anytime.headers = [u'美购ID', u'是否支持随时返现']
    #
    # for key, value in dict_service.items():
    #     row = [key, value]
    #     dt_service_refund_anytime.append(row)
    #
    # open(u'4.xlsx', 'wb').write(dt_service_refund_anytime.xlsx)
    #
    # print(len(dict_service.keys()))


if __name__ == "__main__":
    deaul_refund_anytime()
