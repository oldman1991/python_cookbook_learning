# coding=utf-8
import tablib
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import ExcelWriter


def check_service():
    dict_service = {}
    wb1 = load_workbook(filename='service1.xlsx')
    sheetnames1 = wb1.get_sheet_names()
    if sheetnames1 == 0:
        return
    ws1 = wb1.get_sheet_by_name(sheetnames1[0])

    wb2 = load_workbook(filename='service1.xlsx')
    sheetnames2 = wb2.get_sheet_names()
    if sheetnames2 == 0:
        return
    ws2 = wb2.get_sheet_by_name(sheetnames2[0])

    for rx1 in range(3, ws1.get_highest_row() + 1):
        serviceitem_id_old = int(ws1.cell(row=rx1, column=3).value)
        dict_service[serviceitem_id_old] = {
            '1': ws1.cell(row=rx1, column=1).value,
            '2': ws1.cell(row=rx1, column=2).value,
            '3': ws1.cell(row=rx1, column=3).value,
            '4': ws1.cell(row=rx1, column=4).value,
            '5': ws1.cell(row=rx1, column=5).value,
            '6': ws1.cell(row=rx1, column=6).value,
            '7': ws1.cell(row=rx1, column=7).value,
            '8': ws1.cell(row=rx1, column=8).value,
            '9': ws1.cell(row=rx1, column=9).value,
            '10': ws1.cell(row=rx1, column=10).value,
            '11': ws1.cell(row=rx1, column=11).value
        }
    for rx2 in range(3, ws2.get_highest_row() + 1):
        serviceitem_id = int(ws2.cell(row=rx2, column=3).value)
        gengmei_price = ws2.cell(row=rx2, column=10).value
        pre_payment = ws2.cell(row=rx2, column=11).value
        if not (dict_service[serviceitem_id].get('10') == gengmei_price and dict_service[serviceitem_id].get(
                '11') == pre_payment):
            dict_service[serviceitem_id]['10'] = gengmei_price
            dict_service[serviceitem_id]['11'] = pre_payment

    # wb_result = Workbook()  # 新建一个工作薄
    # ew_result = ExcelWriter(workbook=wb_result)  # 新建一个ExcelWriter,用来写wb_result
    # dest_filename = 'result.xlsx'  # wb_result的名字
    # ws_res = wb_result.worksheets[0]  # 取得wb_result的第一个工作表ws1
    # ws_res.title = 'result'  # 指定ws1的名字
    print('1')

    dt_service_refund_anytime = tablib.Dataset()
    dt_service_refund_anytime.headers = [u'1', u'2', u'3', u'4', u'5', 6,7,8,9,10,11]
    for key, value in dict_service.items():
        row = []
        row.append(dict_service[key]['1'])
        row.append(dict_service[key]['2'])
        row.append(dict_service[key]['3'])
        row.append(dict_service[key]['4'])
        row.append(dict_service[key]['5'])
        row.append(dict_service[key]['6'])
        row.append(dict_service[key]['7'])
        row.append(dict_service[key]['8'])
        row.append(dict_service[key]['9'])
        row.append(dict_service[key]['10'])
        row.append(dict_service[key]['11'])
        dt_service_refund_anytime.append(row)
    print('2')
    open(u'service3.xlsx', 'wb').write(dt_service_refund_anytime.xlsx)
    print('Done')


if __name__ == "__main__":
    check_service()
